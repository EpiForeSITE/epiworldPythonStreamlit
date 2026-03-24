"""
Generic reader for XLSX parameter files. Expects a spreadsheet with at least two columns:

    - Column A: parameter name
    - Column B: parameter value
    - Column C (optional): description or notes, ignored during loading
"""

from io import BytesIO
from typing import IO, Any

import openpyxl
from openpyxl import Workbook
from pydantic import BaseModel

from epicc.formats.base import BaseFormat

# Expected column indices (0-based)
_COL_PARAMETER = 0
_COL_VALUE = 1


class XLSXFormat(BaseFormat[Workbook]):
    """
    Reader for XLSX parameter files.

    Expects a spreadsheet with at least two columns:
        - Column A: parameter name
        - Column B: parameter value
        - Column C (optional): description or notes, ignored during loading

    The first row is treated as a header and skipped. Empty rows are skipped.
    Parameter names may use dot notation to represent nested structure,
    e.g. "costs.latent" will be parsed into {"costs": {"latent": <value>}}.
    """

    def read(self, data: IO) -> tuple[dict[str, Any], Workbook]:
        """
        Read an XLSX file and return its contents as a dictionary.

        Args:
            data: Input stream containing the XLSX data.

        Returns:
            A tuple containing:
              - Dictionary representation of the XLSX contents.
              - Workbook object representing the XLSX file, which can be used as a template for writing.

        Raises:
            FileNotFoundError: If the file does not exist.
            ValueError: If the file cannot be parsed or is missing required columns.
        """

        try:
            wb = openpyxl.load_workbook(data, data_only=False)
        except Exception as e:
            raise ValueError(f"Failed to open XLSX file {self.path}") from e

        ws = wb.active
        if not ws or ws.max_column < 2:
            raise ValueError(
                f"XLSX file {self.path} must have at least 2 columns for parameters and values."
            )

        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 2:
            raise ValueError(
                f"XLSX file {self.path} must have a header row and at least one data row."
            )

        opaque: dict[str, Any] = {}

        for row in rows[1:]:  # skip header
            if not row or row[_COL_PARAMETER] is None:
                continue

            key = str(row[_COL_PARAMETER]).strip()
            value = row[_COL_VALUE] if len(row) > _COL_VALUE else None

            if not key:
                continue

            _set_nested(opaque, key, value)

        return opaque, wb

    def write(self, data: dict[str, Any], template: Workbook | None = None) -> bytes:
        """
        Write a dictionary to an XLSX file.

        Args:
            data: Dictionary to write.
            template: Optional Workbook object to use as a template for writing. If provided, the
                      structure and formatting of the template will be preserved as much as possible.

        Returns:
            Byte array containing the XLSX data.
        """

        wb = template or Workbook()
        ws = wb.active
        assert ws is not None, "Workbook must have an active worksheet (bug)."

        # Populate with the provided data piecewise.
        for row in ws.iter_rows():
            key_cell = row[_COL_PARAMETER]
            val_cell = row[_COL_VALUE]
            if key_cell.value in data:
                val_cell.value = data[key_cell.value]  # type: ignore

        # Awful, but openpyxl is only capable of doing it this way.
        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    def write_template(self, model: BaseModel) -> bytes:
        """Write an XLSX template from a model instance.

        Produces a three-column spreadsheet (Parameter, Value, Description).
        Nested models are flattened to dot-notation keys. Descriptions come
        from ``Field(description=...)`` on each field.
        """
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.append(["Parameter", "Value", "Description"])
        for key, value, description in _flatten(model):
            ws.append([key, value, description])
        output = BytesIO()
        wb.save(output)
        return output.getvalue()


def _flatten(
    model: BaseModel, prefix: str = ""
) -> list[tuple[str, Any, str]]:
    """Recursively flatten a model instance to ``[(dot_key, value, description)]``."""
    rows: list[tuple[str, Any, str]] = []
    for name, field_info in type(model).model_fields.items():
        key = f"{prefix}.{name}" if prefix else name
        value = getattr(model, name)
        description = field_info.description or ""
        if isinstance(value, BaseModel):
            rows.extend(_flatten(value, prefix=key))
        else:
            rows.append((key, value, description))
    return rows


def _set_nested(d: dict, key: str, value: Any) -> None:
    """
    Set a value in a nested dictionary using dot-notation keys.

    For example, "costs.latent" with value 300.0 produces:
        {"costs": {"latent": 300.0}}

    Args:
        d: Dictionary to mutate.
        key: Dot-separated key string.
        value: Value to set.
    """

    parts = key.split(".")
    for part in parts[:-1]:
        d = d.setdefault(part, {})

    d[parts[-1]] = value


__all__ = ["XLSXFormat"]

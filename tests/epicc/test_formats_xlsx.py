"""Tests for epicc.formats.xlsx (XLSXFormat)."""

from io import BytesIO

import openpyxl
import pytest

from epicc.formats.xlsx import XLSXFormat


def _make_xlsx(rows: list[list]) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _fmt() -> XLSXFormat:
    return XLSXFormat("test.xlsx")


def test_read_dot_notation_creates_nested():
    buf = _make_xlsx([
        ["param", "value"],
        ["costs.latent", 300],
        ["costs.active", 500],
    ])
    data, _ = _fmt().read(buf)
    assert data == {"costs": {"latent": 300, "active": 500}}


def test_read_skips_empty_rows():
    buf = _make_xlsx([["param", "value"], ["a", 1], [None, None], ["b", 2]])
    data, _ = _fmt().read(buf)
    assert data == {"a": 1, "b": 2}


def test_read_too_few_columns_raises():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["only_one_col"])
    ws.append(["value"])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    with pytest.raises(ValueError, match="at least 2 columns"):
        _fmt().read(buf)


def test_write_with_template_round_trips():
    buf = _make_xlsx([["param", "value"], ["x", 1], ["y", 2]])
    _, wb_template = _fmt().read(buf)

    result_bytes = _fmt().write({"x": 99}, wb_template)

    wb_out = openpyxl.load_workbook(BytesIO(result_bytes))
    rows = list(wb_out.active.iter_rows(values_only=True))
    row_dict = {r[0]: r[1] for r in rows[1:] if r[0] is not None}
    assert row_dict["x"] == 99
